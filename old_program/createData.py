import csv
import json
import statistics
import math
import pandas as pd
import openpyxl as px
import matplotlib.pyplot as plt
import numpy as np
from scipy.stats import norm
from decimal import Decimal, ROUND_HALF_UP, ROUND_HALF_EVEN


def createInitialData():

    play_dictList = get_playDictList()

    action = [ [], [] ]
    get_yard, possibilities, p_prob = [], [], []

    wb = px.load_workbook("/Users/shimano/webappSample/GameTheory/Data/Create/create2017xl.xlsx")
    ws = wb["Sheet1"]

    for i in range(50):
        if ws.cell(row=2+i, column=1).value == None:
            break
        action[0].append(i)
        get_yard.append([])
        possibilities.append([])
        p_prob.append([])
        for j in range(50):
            if ws.cell(row=1, column=2+j).value == None:
                break
            if j not in action[1]:
                action[1].append(j)
            for k in play_dictList:
                if ws.cell(row=2+i, column=1).value in k and ws.cell(row=1, column=2+j).value in k:
                    #確率関数作成
                    Min = math.floor(norm.ppf(0.00001, statistics.mean(play_dictList[k]), statistics.stdev(play_dictList[k])))
                    if Min < -10:
                        Min = -100
                    Max = math.floor(norm.ppf(0.99999, statistics.mean(play_dictList[k]), statistics.stdev(play_dictList[k])))
                    if Max > 100:
                        Max = 100
                    yard_and_prob, def_yard_clum, def_possib_clum, def_prob_clum = {}, [], [], []
                    itenum = 0
                    while Min < Max:
                        p1 = norm.cdf(Min-0.5, statistics.mean(play_dictList[k]), statistics.stdev(play_dictList[k]))
                        p2 = norm.cdf(Min+0.5, statistics.mean(play_dictList[k]), statistics.stdev(play_dictList[k]))
                        yard_and_prob[str(Min)] = Decimal(p2-p1)
                        def_yard_clum.append(Min)
                        if len(str(itenum)) == 1:
                            def_possib_clum.append("00"+str(itenum))
                        elif len(str(itenum)) == 2:
                            def_possib_clum.append("0"+str(itenum))
                        else:
                            def_possib_clum.append(str(itenum))
                        def_prob_clum.append(p2-p1)
                        itenum += 1
                        Min += 1
                    get_yard[i].append(def_yard_clum)
                    possibilities[i].append(def_possib_clum)
                    p_prob[i].append(def_prob_clum)

    return action, get_yard, possibilities, p_prob


def get_playDictList():

    plays = open("/Users/shimano/webappSample/GameTheory/Data/Origin/plays.csv", "r")

    p_reader = csv.reader(plays)
    p_header = next(p_reader)

    data = []
    play_dict, num_dict, play_dictList = {}, {}, {}
    for i, row in enumerate(p_reader):
        if row[19] == "FALSE" and row[18] == "FALSE":
            if str(row[9]) != "NA" and str(row[10]) != "NA" and str(row[11]) != "NA" and str(row[13]) != "NA":
                if "WR" not in str(row[13]) and "OL" not in str(row[13]) and "TE" not in str(row[13]):
                    if "pass" in str(row[26]):
                        hoge = {
                            "Down": str(row[4]),
                            "YardLineNumber": str(row[8]),
                            "OffenseFormation": str(row[9]),
                            "OffensePersonnel": str(row[10]),
                            "DefenseInTheBox": str(row[11]),
                            "DefensePersonnel": str(row[13]),
                            "PlayType": "Pass",
                            "RunGap": "NA",
                            "PassLength": "short" if "short" in str(row[26]) else "deep",
                            "ActionSide": "middle" if "middle" in str(row[26]) else "right" if "right" in str(row[26]) else "left",
                            "PassYards": str(row[22]),
                            "PassResult": str(row[23]),
                            "YardsAfterCatch": str(row[24]),
                            "PlayResult": str(row[25]),
                            "Description": str(row[26]),
                        }
                        off_name = hoge["OffenseFormation"]+"/"+hoge["OffensePersonnel"]+"/"+hoge["PlayType"]+"/"+hoge["PassLength"]+"/"+hoge["ActionSide"]
                        def_name = hoge["DefenseInTheBox"]+"/"+hoge["DefensePersonnel"]
                        action_profile = off_name+" : "+def_name
                        if action_profile not in play_dict:
                            play_dict[action_profile] = int(hoge["PlayResult"])
                            num_dict[action_profile]  = [off_name, def_name, 1]#1
                            play_dictList[action_profile] = [int(hoge["PlayResult"])]
                        else:
                            play_dict[action_profile] += int(hoge["PlayResult"])
                            num_dict[action_profile][2]  += 1
                            play_dictList[action_profile].append(int(hoge["PlayResult"]))
                    else:
                        hoge = {
                            "Down": str(row[4]),
                            "YardLineNumber": str(row[8]),
                            "OffenseFormation": str(row[9]),
                            "OffensePersonnel": str(row[10]),
                            "DefenseInTheBox": str(row[11]),
                            "DefensePersonnel": str(row[13]),
                            "PlayType": "Run",
                            "RunGap": "end" if "end" in str(row[26]) else "guard" if "guard" in str(row[26]) else "tackle" if "tackle" in str(row[26]) else "NA",
                            "PassLength": "NA",
                            "ActionSide": "left" if "left" in str(row[26]) else "right" if "right" in str(row[26]) else "middle",
                            "PassYards": str(row[22]),
                            "PassResult": str(row[23]),
                            "YardsAfterCatch": str(row[24]),
                            "PlayResult": str(row[25]),
                            "Description": str(row[26]),
                        }
                        off_name = hoge["OffenseFormation"]+"/"+hoge["OffensePersonnel"]+"/"+hoge["PlayType"]+"/"+hoge["RunGap"]+"/"+hoge["ActionSide"]
                        def_name = hoge["DefenseInTheBox"]+"/"+hoge["DefensePersonnel"]
                        action_profile = off_name+" : "+def_name
                        if action_profile not in play_dict:
                            play_dict[action_profile] = int(hoge["PlayResult"])
                            num_dict[action_profile]  = [off_name, def_name, 1]#1
                            play_dictList[action_profile] = [int(hoge["PlayResult"])]
                        else:
                            play_dict[action_profile] += int(hoge["PlayResult"])
                            num_dict[action_profile][2]  += 1
                            play_dictList[action_profile].append(int(hoge["PlayResult"]))
                    
                    data.append(hoge)
    
    return play_dictList