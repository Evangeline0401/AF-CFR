import csv
import json
import statistics
import math
import pandas as pd
import openpyxl as px
import matplotlib.pyplot as plt
import numpy as np
from scipy.stats import norm
from decimal import Decimal, ROUND_HALF_UP, ROUND_HALF_EVEN


plays = open("Origin/plays.csv", "r")

p_reader = csv.reader(plays)
p_header = next(p_reader)

data, off_list, def_list = [], [], []
play_dict, num_dict, play_dictList, play_distdict = {}, {}, {}, {}

for i, row in enumerate(p_reader):
    if row[19] == "FALSE" and row[18] == "FALSE":
        if str(row[9]) != "NA" and str(row[10]) != "NA" and str(row[11]) != "NA" and str(row[13]) != "NA":
            if "WR" not in str(row[13]) and "OL" not in str(row[13]) and "TE" not in str(row[13]):
                if "pass" in str(row[26]):
                    hoge = {
                        "Down": str(row[4]),
                        "YardLineNumber": str(row[8]),
                        "OffenseFormation": str(row[9]),
                        "OffensePersonnel": str(row[10]),
                        "DefenseInTheBox": str(row[11]),
                        "DefensePersonnel": str(row[13]),
                        "PlayType": "Pass",
                        "RunGap": "NA",
                        "PassLength": "short" if "short" in str(row[26]) else "deep",
                        "ActionSide": "middle" if "middle" in str(row[26]) else "right" if "right" in str(row[26]) else "left",
                        "PassYards": str(row[22]),
                        "PassResult": str(row[23]),
                        "YardsAfterCatch": str(row[24]),
                        "PlayResult": str(row[25]),
                        "Description": str(row[26]),
                    }
                    off_name = hoge["OffenseFormation"]+"/"+hoge["OffensePersonnel"]+"/"+hoge["PlayType"]+"/"+hoge["PassLength"]+"/"+hoge["ActionSide"]
                    def_name = hoge["DefenseInTheBox"]+"/"+hoge["DefensePersonnel"]
                    action_profile = off_name+" : "+def_name
                    if action_profile not in play_dict:
                        play_dict[action_profile] = int(hoge["PlayResult"])
                        num_dict[action_profile]  = [off_name, def_name, 1]#1
                        play_dictList[action_profile] = [int(hoge["PlayResult"])]
                        play_distdict[action_profile] = {str(hoge["PlayResult"]): 1}
                    else:
                        play_dict[action_profile] += int(hoge["PlayResult"])
                        num_dict[action_profile][2]  += 1
                        play_dictList[action_profile].append(int(hoge["PlayResult"]))
                        if str(hoge["PlayResult"]) not in play_distdict[action_profile]:
                            play_distdict[action_profile][str(hoge["PlayResult"])] = 1
                        else:
                            play_distdict[action_profile][str(hoge["PlayResult"])] += 1
                else:
                    hoge = {
                        "Down": str(row[4]),
                        "YardLineNumber": str(row[8]),
                        "OffenseFormation": str(row[9]),
                        "OffensePersonnel": str(row[10]),
                        "DefenseInTheBox": str(row[11]),
                        "DefensePersonnel": str(row[13]),
                        "PlayType": "Run",
                        "RunGap": "end" if "end" in str(row[26]) else "guard" if "guard" in str(row[26]) else "tackle" if "tackle" in str(row[26]) else "NA",
                        "PassLength": "NA",
                        "ActionSide": "left" if "left" in str(row[26]) else "right" if "right" in str(row[26]) else "middle",
                        "PassYards": str(row[22]),
                        "PassResult": str(row[23]),
                        "YardsAfterCatch": str(row[24]),
                        "PlayResult": str(row[25]),
                        "Description": str(row[26]),
                    }
                    off_name = hoge["OffenseFormation"]+"/"+hoge["OffensePersonnel"]+"/"+hoge["PlayType"]+"/"+hoge["RunGap"]+"/"+hoge["ActionSide"]
                    def_name = hoge["DefenseInTheBox"]+"/"+hoge["DefensePersonnel"]
                    action_profile = off_name+" : "+def_name
                    if action_profile not in play_dict:
                        play_dict[action_profile] = int(hoge["PlayResult"])
                        num_dict[action_profile]  = [off_name, def_name, 1]#1
                        play_dictList[action_profile] = [int(hoge["PlayResult"])]
                        play_distdict[action_profile] = {str(hoge["PlayResult"]): 1}
                    else:
                        play_dict[action_profile] += int(hoge["PlayResult"])
                        num_dict[action_profile][2]  += 1
                        play_dictList[action_profile].append(int(hoge["PlayResult"]))
                        if str(hoge["PlayResult"]) not in play_distdict[action_profile]:
                            play_distdict[action_profile][str(hoge["PlayResult"])] = 1
                        else:
                            play_distdict[action_profile][str(hoge["PlayResult"])] += 1
                
                data.append(hoge)


j = open("/Users/shimano/webappSample/GameTheory/Data/Create/data2017.json", "w")
json.dump(data, j, indent=3)

plays.close()




num_dict = sorted(num_dict.items(), key=lambda x: x[1][2], reverse=True)
for i in num_dict:
    if i[1][0] not in off_list:
        off_list.append(i[1][0])
    if i[1][1] not in def_list:
        def_list.append(i[1][1])
"""
#ここは必要なときにコメントアウト外す
wb = px.Workbook()
ws = wb.active
ws = wb.create_sheet(0)
for i, v in enumerate(off_list):
    ws.cell(row=2+i, column=1).value = v
for i, v in enumerate(def_list):
    ws.cell(row=1, column=2+i).value = v

for i in range(len(off_list)):
    for j in range(len(def_list)):
        for k in num_dict:
            if ws.cell(row=2+i, column=1).value in k[0] and ws.cell(row=1, column=2+j).value in k[0]:
                ws.cell(row=2+i, column=2+j).value = k[1][2]
                break

wb.save("Create/shimano.xlsx")
"""

X = np.arange(-100, 100, 0.1)

#Average
wb = px.load_workbook("Create/create2017xl.xlsx")
ws = wb["Sheet1"]

for i in range(50):
    if ws.cell(row=2+i, column=1).value == None:
        break
    for j in range(50):
        if ws.cell(row=1, column=2+j).value == None:
            break
        for k in play_dictList:
            if ws.cell(row=2+i, column=1).value in k and ws.cell(row=1, column=2+j).value in k:
                ws.cell(row=2+i, column=2+j).value = statistics.mean(play_dictList[k])#play_dict[k] / ws.cell(row=2+i, column=2+j).value

wb.save("Create/create2017xl_ave.xlsx")


#standard deviation
wb = px.load_workbook("Create/create2017xl.xlsx")
ws = wb["Sheet1"]

for i in range(50):
    if ws.cell(row=2+i, column=1).value == None:
        break
    for j in range(50):
        if ws.cell(row=1, column=2+j).value == None:
            break
        for k in play_dictList:
            if ws.cell(row=2+i, column=1).value in k and ws.cell(row=1, column=2+j).value in k:
                ws.cell(row=2+i, column=2+j).value = statistics.stdev(play_dictList[k])

                #正規分布のPNG作成
                Y = norm.pdf(X, statistics.mean(play_dictList[k]), statistics.stdev(play_dictList[k]))
                fig = plt.figure()
                plt.plot(X, Y ,color='r')
                plt.title(k)
                plt.xlim(norm.ppf(0.00001, statistics.mean(play_dictList[k]), statistics.stdev(play_dictList[k])),
                         norm.ppf(0.99999, statistics.mean(play_dictList[k]), statistics.stdev(play_dictList[k])))
                fig.savefig("Create/PNG/SD/"+ str(i) +"-"+ str(j) +".png")
                plt.close()

                #確率関数作成
                Min = math.floor(norm.ppf(0.00001, statistics.mean(play_dictList[k]), statistics.stdev(play_dictList[k])))
                if Min < -100:
                    Min = -100
                Max = math.floor(norm.ppf(0.99999, statistics.mean(play_dictList[k]), statistics.stdev(play_dictList[k])))
                if Max > 100:
                    Max = 100
                yard_and_prob = {}
                while Min < Max:
                    p1 = norm.cdf(Min-0.5, statistics.mean(play_dictList[k]), statistics.stdev(play_dictList[k]))
                    p2 = norm.cdf(Min+0.5, statistics.mean(play_dictList[k]), statistics.stdev(play_dictList[k]))
                    yard_and_prob[str(Min)] = Decimal(p2-p1)
                    Min += 1

wb.save("Create/create2017xl_sd.xlsx")



#Distribution
wb = px.load_workbook("Create/create2017xl.xlsx")
ws = wb["Sheet1"]

for i in range(50):
    if ws.cell(row=2+i, column=1).value == None:
        break
    for j in range(50):
        if ws.cell(row=1, column=2+j).value == None:
            break
        for k in play_distdict:
            if ws.cell(row=2+i, column=1).value in k and ws.cell(row=1, column=2+j).value in k:
                s_play_distdict = sorted(play_distdict[k].items(), key=lambda x: int(x[0]))
                yard, num = [], []
                for dist in s_play_distdict:
                    yard.append(dist[0])
                    num.append(dist[1])
                fig = plt.figure()
                plt.bar(yard, num)
                #plt.xlim(s_play_distdict[0][0], s_play_distdict[len(s_play_distdict)-1][0])
                plt.title(k)
                plt.xlabel("Yard")
                plt.ylabel("Num")
                fig.savefig("Create/PNG/Dist/"+ str(i) +"-"+ str(j) +".png")
                plt.close()
